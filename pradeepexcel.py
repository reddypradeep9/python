import openpyxl,os,sys
#print(os.getcwd())
os.chdir('C:\\Users\\PradeepReddyS\\Downloads')
wb=openpyxl.load_workbook('VF_IT.xlsx')
#print(wb.sheetnames)
sheet=wb['Package Pricepoints']
row=sheet.max_row
col=sheet.max_column
r=[]
c=[]

pk_name=input('Enter the packag name which you want to search')


for i in range(1,row+1):
    for j in range(1,col+1):
        if sheet.cell(row=i,column=j).value==pk_name:

            r.append(sheet.cell(row=i,column=j).row)
            c.append(sheet.cell(row=i,column=j).column)



print(r)
print(c)
for i in range(3,27):
    print(str(sheet.cell(row=3,column=i).value).ljust(30).upper(),end='\t')

print()
print()
for ro in r:
    for co in range(c[0],c[0]+24):
        print(str(sheet.cell(row=ro,column=co).value).ljust(30),end='\t')
    print()
