import openpyxl,os
from openpyxl.styles import Font
os.chdir('C:\\Users\\PradeepReddyS\\Downloads')
wb=openpyxl.load_workbook('VF_IT.xlsx')
#print(wb.sheetnames)
sheet=wb['Package Pricepoints']
print(sheet['A14'].value)
#print(sheet.cell(row=1,column=224).value)

for i in range(50,65):
    print(sheet.cell(row=i,column=6).value)
print(sheet.max_row)                #for getting maximum row
print(sheet.max_column)             #for getting maximum col
print(openpyxl.cell.get_column_letter(1))
print(openpyxl.cell.column_index_from_string('AA'))
wb.create_sheet(title='pradeep',index=0)
sheet.row_dimensions[1].height=70
sheet.column_dimensions['A']=122            #Change dimensions of rows and col
sheet['B1'].font=Font(sz=14,bold=True,italic=True)




wb.save('VF_IT.xlsx')








